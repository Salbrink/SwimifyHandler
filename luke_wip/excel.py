from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment


class ExcelSheet:

    def __init__(self) -> None:
        self.workbook = load_workbook("Template.xlsx")
        self.dst_sheet = self.workbook["Empty"]
        self.row_index = self.dst_sheet.max_row

    def save_one_swimmer(self, sheet, name, pb_sc, pb_lc, event):
        src_sheet = self.workbook[sheet]

        # Remember the first row to get the merge correct
        row_index = self.row_index + 2

        # To add some spacing between every event
        self.row_index += 1

        # Loop through the rows and cells in the source sheet
        for row in src_sheet.iter_rows():
            self.row_index += 1
            for cell in row:
                # Write the cell value in the destination sheet keeping the source formatting
                id = cell.column_letter + str(self.row_index)
                self.dst_sheet[id].font = copy(cell.font)
                self.dst_sheet[id].alignment = copy(cell.alignment)
                self.dst_sheet[id].number_format = copy(cell.number_format)
                self.dst_sheet[id].protection = copy(cell.protection)
                self.dst_sheet[id].fill = copy(cell.fill)
                self.dst_sheet[id].border = copy(cell.border)

                # Update cell value if key words are found
                if str(cell.value) == "$name$":
                    new_cell_value = name
                elif str(cell.value) == "$pb_sc$":
                    new_cell_value = pb_sc
                elif str(cell.value) == "$pb_lc$":
                    new_cell_value = pb_lc
                elif str(cell.value) == "$event$":
                    new_cell_value = event
                else:
                    new_cell_value = cell.value

                self.dst_sheet[id].value = new_cell_value

                # Cell width and height need to handled seperately
                self.dst_sheet.column_dimensions[cell.column_letter].width = (
                    src_sheet.column_dimensions[cell.column_letter].width
                )
                self.dst_sheet.row_dimensions[self.row_index].height = (
                    src_sheet.row_dimensions[cell.row].height
                )

        # Do some merging to get the name and event nicely looking
        name_range_start = "B" + str(row_index)
        name_range = name_range_start + ":C" + str(row_index)
        event_range_start = "B" + str(row_index + 1)
        event_range = event_range_start + ":H" + str(row_index + 1)

        self.dst_sheet.merge_cells(name_range)
        self.dst_sheet.merge_cells(event_range)

        self.dst_sheet[name_range_start].alignment = Alignment(horizontal="center")
        self.dst_sheet[event_range_start].alignment = Alignment(horizontal="center")

        self.workbook.save("test.xlsx")
